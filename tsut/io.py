import ast
import copy
import json
from openpyxl import Workbook, load_workbook

from .api import UsersAndGroups, User, Group, eprint

"""
Copyright 2018 ThoughtSpot
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in all copies or substantial portions
of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

# -------------------------------------------------------------------------------------------------------------------

"""Classes to read and write users and groups."""


class UGXLSWriter:
    """
    Writes users and groups to an Excel spreadsheet.
    """

    def write(self, users_and_groups, filename):
        """
        Writes the content to the given file.
        :param users_and_groups:  The UsersAndGroups object to write.
        :type users_and_groups: UsersAndGroups
        :param filename:  Name of the file to write to.  No extension is expected and one will be added.
        :type filename: str
        """
        workbook = Workbook()
        workbook.remove(
            workbook.active
        )  # remove the default sheet since we'll be creating the ones we want.
        self._write_users(workbook, users_and_groups.get_users())
        self._write_groups(workbook, users_and_groups.get_groups())
        if not (filename.endswith("xls") or filename.endswith("xlsx")):
            filename += ".xlsx"

        workbook.save(filename)

    def _write_users(self, workbook, users):
        """
        Writes the users to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param users:  The list of groups to write.
        :type users: list of User
        :return:
        """
        ws = workbook.create_sheet(title="Users")
        self._write_header(
            ws,
            [
                "Name",
                "Password",
                "Display Name",
                "Email",
                "Groups",
                "Visibility"
            ],
        )
        cnt = 2  # start after header.
        for user in users:
            ws.cell(column=1, row=cnt, value=user.name)
            ws.cell(column=2, row=cnt, value=user.password)
            ws.cell(column=3, row=cnt, value=user.displayName)
            ws.cell(column=4, row=cnt, value=user.mail)
            ws.cell(column=5, row=cnt, value=json.dumps(user.groupNames))
            ws.cell(column=6, row=cnt, value=user.visibility)
            cnt += 1

    def _write_groups(self, workbook, groups):
        """
        Writes the groups to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param groups:  The list of groups to write.
        :type groups: list
        :return:
        """
        ws = workbook.create_sheet(title="Groups")
        self._write_header(
            ws,
            [
                "Name",
                "Display Name",
                "Description",
                "Groups",
                "Visibility",
                "Privileges",
            ],
        )
        cnt = 2  # start after header.
        for group in groups:
            ws.cell(column=1, row=cnt, value=group.name)
            ws.cell(column=2, row=cnt, value=group.displayName)
            ws.cell(column=3, row=cnt, value=group.description)
            ws.cell(column=4, row=cnt, value=json.dumps(group.groupNames))
            ws.cell(column=5, row=cnt, value=group.visibility)
            privileges = group.privileges if group.privileges else []
            ws.cell(column=6, row=cnt, value=json.dumps(privileges))
            cnt += 1

    @staticmethod
    def _write_header(worksheet, cols):
        """
        Writes the header for the given worksheet in row 1.
        :param worksheet:  Worksheet to write to.
        :param cols:  List of columns to write.
        """
        for ccnt in range(0, len(cols)):
            worksheet.cell(column=(ccnt + 1), row=1, value=cols[ccnt])


class UGXLSReader:
    """
    Reads user and group info from an Excel file that is formatted the same as the UGXLSWriter writes.
    """

    required_sheets = ["Users", "Groups"]
    required_columns = {
        "Users": [
            "Name",
            "Password",
            "Display Name",
            "Email",
            "Groups",
            "Visibility",
        ],
        "Groups": [
            "Name",
            "Display Name",
            "Description",
            "Groups",
            "Visibility",
            "Privileges"
        ],
    }

    def __init__(self):
        """
        Creates a new UGXLSReader
        """
        self.workbook = None
        self.indices = {}
        self.users_and_groups = UsersAndGroups()

    def read_from_excel(self, filepath):
        """
        Reads users and groups from the given file.
        :param filepath:  Path to the Excel file to read from.
        :type filepath: str
        :return: Returns the users and groups read from the Excel file.  The users and groups are not validated
        :rtype UsersAndGroups
        so that they can be modified prior to validation.
        """
        self.workbook = load_workbook(filename=filepath, read_only=True)
        if self._verify_file_format():
            self._get_column_indices()
            self._read_users_from_workbook()
            self._read_groups_from_workbook()
        return self.users_and_groups

    def _verify_file_format(self):
        """
        :return: True if the format of the workbook is valid.
        :rtype: bool
        """
        is_valid = True
        sheet_names = self.workbook.sheetnames
        for required_sheet in UGXLSReader.required_sheets:
            if required_sheet not in sheet_names:
                eprint("Error:  missing sheet %s!" % required_sheet)
                is_valid = False
            else:
                sheet = self.workbook[required_sheet]
                header_row = [cell.value for cell in list(self.workbook[required_sheet].iter_rows(max_row=1))[0]]

                for required_column in UGXLSReader.required_columns[
                    required_sheet
                ]:
                    if required_column not in header_row:
                        eprint(
                            "Error:  missing column %s in sheet %s!"
                            % (required_column, required_sheet)
                        )
                        is_valid = False

        return is_valid

    def _get_column_indices(self):
        """
        Reads the sheets to get all of the column indices.  Assumes the format was already checked.
        """
        sheet_names = self.workbook.sheetnames
        for sheet_name in sheet_names:
            if sheet_name in self.required_sheets:
                sheet = self.workbook[sheet_name]
                col_indices = {}
                ccnt = 0
                row = [cell.value for cell in list(self.workbook[sheet_name].iter_rows(min_row=1, max_row=2))[0]]
                for col in row:
                    col_indices[col] = ccnt
                    ccnt += 1
                self.indices[sheet_name] = col_indices

    def _read_users_from_workbook(self):
        """
        Reads all the users from the workbook.
        """

        table_sheet = self.workbook["Users"]
        indices = self.indices["Users"]

        header_row = True
        for row in table_sheet.values:
            values = list(row)
            if header_row:  # have to skip the header row.
                header_row = False
                continue

            # "Name", "Password", "Display Name", "Email", "Description", "Groups", "Visibility"
            username = values[indices["Name"]]
            password = values[indices["Password"]]
            display_name = values[indices["Display Name"]]
            email = values[indices["Email"]]
            groups = []
            if values[indices["Groups"]] and values[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    values[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            visibility = values[indices["Visibility"]]

            try:
                user = User(
                    name=username,
                    password=password,
                    display_name=display_name,
                    mail=email,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one user per line.
                self.users_and_groups.add_user(
                    user, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except:
                eprint(f"Error reading user with name {username}")

    def _read_groups_from_workbook(self):
        """
        Reads all the groups from the workbook.
        """

        table_sheet = self.workbook["Groups"]
        indices = self.indices["Groups"]

        header_row = True
        for row in table_sheet.values:
            values = list(row)
            if header_row:  # have to skip the header row.
                header_row = False
                continue


            # Name", "Display Name", "Description", "Groups", "Visibility", "Privileges"
            group_name = values[indices["Name"]]
            display_name = values[indices["Display Name"]]
            description = values[indices["Description"]]
            visibility = values[indices["Visibility"]]

            groups = []
            if values[indices["Groups"]] and values[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    values[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]

            privileges = []
            if values[indices["Privileges"]] and values[
                indices["Privileges"]
            ]:
                privileges = ast.literal_eval(
                    values[indices["Privileges"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]

            try:
                group = Group(
                    name=group_name,
                    display_name=display_name,
                    description=description,
                    group_names=groups,
                    visibility=visibility,
                    privileges=privileges
                )
                # The format should be consistent with only one group per line.
                self.users_and_groups.add_group(
                    group, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except Exception:
                eprint("Error reading group with name %s" % group_name)


class UGCSVReader:
    """
    Reads users and groups from CSV.  All users and groups are in a single file.
    """
    DEFAULT_GROUP_FIELD_MAPPING = {
        "name": "Group Name",
        "display_name": "Group Display Name",
        "description": "Group Description",
        "group_names": "Group Names",
        "visibility": "Group Visibility",
        "privileges": "Group Privileges"
    }
    DEFAULT_USER_FIELD_MAPPING = {
        "name": "User Name",
        "display_name": "User Display Name",
        "password": "User Password",
        "description": "User Description",
        "group_names": "User Group Names",
        "visibility": "User Visibility"
    }

    def __init__(self,
                 user_field_mapping=DEFAULT_USER_FIELD_MAPPING,
                 group_field_mapping=DEFAULT_GROUP_FIELD_MAPPING,
                 delimiter=","):
        """
        Creates a new CSV reader that can read based on the field mapping and delimiter.  While this class can
        cause groups to be created, the primary use is to have groups that will be
        :param user_field_mapping: The mapping of columns to values for users.
        :type user_field_mapping: dict of str:str
        :param group_field_mapping: The mapping of columns to values for groups.
        :type group_field_mapping: dict of str:str
        :param delimiter: The delimiter to use.
        """
        self.user_field_mapping = copy.copy(user_field_mapping)
        self.group_field_mapping = copy.copy(group_field_mapping)
        self.delimiter = delimiter

        self.validate_fields()

    def validate_fields(self):
        """
        Verifies that the minimal required field mappings exist.  Raises a ValueError if not.
        :return: None
        :raises: ValueError
        """
        if "name" not in self.user_field_mapping.keys():
            raise ValueError("Missing name parameter for users.")
        if "name" not in self.group_field_mapping.keys():
            raise ValueError("Missing name parameter for groups.")

    def read_from_file(self, user_file, group_file=None):
        """
        Loads users and groups from the files.  If the group_file is not provided, the groups will be created from the
        user file with just the names.
        :param user_file: Path to the user file to read from.
        :type user_file: str
        :param group_file: Path to the group file to read from.
        :type group_file: str
        :return: Users and groups object.
        :rtype: UsersAndGroups
        """
        pass


